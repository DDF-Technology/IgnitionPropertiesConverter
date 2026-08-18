import tempfile
import unittest
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook

from converter import PropertiesFormatError, export_to_excel, import_from_excel, read_properties_file


class ConverterTests(unittest.TestCase):
    def test_properties_parser_supports_escapes_unicode_and_continuations(self):
        with tempfile.TemporaryDirectory() as directory:
            source = Path(directory) / "Language_it.properties"
            source.write_text("# commento\nmenu\\:title=Caff\\u00e8\nmultiline=prima\\\n  seconda\n", encoding="utf-8")

            self.assertEqual({"menu:title": "Caffè", "multiline": "primaseconda"}, read_properties_file(source))

    def test_duplicate_properties_key_is_rejected(self):
        with tempfile.TemporaryDirectory() as directory:
            source = Path(directory) / "Language_it.properties"
            source.write_text("same=uno\nsame=due\n", encoding="utf-8")

            with self.assertRaisesRegex(PropertiesFormatError, "duplicata"):
                read_properties_file(source)

    def test_excel_round_trip_preserves_special_characters(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            italian = root / "Language_it.properties"
            english = root / "Language_en.properties"
            workbook = root / "translations.xlsx"
            output = root / "output"
            italian.write_text("menu\\:title=Caffè \\#1\npath=C\\\\Programmi\n", encoding="utf-8")
            english.write_text("menu\\:title=Coffee \\#1\npath=C\\\\Programs\n", encoding="utf-8")

            export_to_excel([italian, english], workbook)
            import_from_excel(workbook, output)

            self.assertEqual(read_properties_file(italian), read_properties_file(output / "Language_it.properties"))
            self.assertEqual(read_properties_file(english), read_properties_file(output / "Language_en.properties"))

    def test_invalid_language_column_is_rejected(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            workbook = root / "translations.xlsx"
            pd.DataFrame({"key": ["welcome"], "../escape": ["no"]}).to_excel(workbook, index=False)

            with self.assertRaisesRegex(PropertiesFormatError, "non validi"):
                import_from_excel(workbook, root / "output")

    def test_export_preserves_value_starting_with_equals_as_text(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            source = root / "Language_it.properties"
            workbook_path = root / "translations.xlsx"
            source.write_text("formula==SUM(A1:A2)\nliteral\\=key==valore\n", encoding="utf-8")

            export_to_excel([source], workbook_path)

            workbook = load_workbook(workbook_path, data_only=False)
            try:
                values = {row[0].value: row[1] for row in workbook.active.iter_rows(min_row=2)}
                self.assertEqual("=SUM(A1:A2)", values["formula"].value)
                self.assertEqual("s", values["formula"].data_type)
                self.assertEqual("=valore", values["literal=key"].value)
                self.assertEqual("s", values["literal=key"].data_type)
            finally:
                workbook.close()

    def test_import_rejects_excel_formulas(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            workbook_path = root / "translations.xlsx"
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.append(["key", "it"])
            worksheet.append(["danger", "=HYPERLINK(\"https://example.invalid\",\"test\")"])
            workbook.save(workbook_path)

            with self.assertRaisesRegex(PropertiesFormatError, "contiene formule"):
                import_from_excel(workbook_path, root / "output")

    def test_empty_selection_and_duplicate_language_are_rejected(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            with self.assertRaisesRegex(PropertiesFormatError, "almeno un file"):
                export_to_excel([], root / "translations.xlsx")

            first = root / "Language_it.properties"
            second = root / "Other_IT.properties"
            first.write_text("welcome=Ciao\n", encoding="utf-8")
            second.write_text("welcome=Salve\n", encoding="utf-8")
            with self.assertRaisesRegex(PropertiesFormatError, "lingua duplicato"):
                export_to_excel([first, second], root / "translations.xlsx")


if __name__ == "__main__":
    unittest.main()

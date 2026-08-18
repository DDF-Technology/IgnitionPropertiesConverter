"""Conversione sicura tra file Java/Ignition ``.properties`` e workbook Excel."""

from __future__ import annotations

import os
import re
import tempfile
from collections.abc import Iterable
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


class PropertiesFormatError(ValueError):
    """Segnala contenuti che causerebbero una conversione ambigua o distruttiva."""


_LANGUAGE_PATTERN = re.compile(r"^[A-Za-z][A-Za-z0-9_-]*$")


def _logical_lines(lines: Iterable[str]) -> Iterable[tuple[int, str]]:
    buffer = ""
    start_line = 0
    for line_number, physical_line in enumerate(lines, 1):
        line = physical_line.rstrip("\r\n")
        if not buffer:
            start_line = line_number
        else:
            line = line.lstrip(" \t\f")
        buffer += line
        trailing_backslashes = len(buffer) - len(buffer.rstrip("\\"))
        if trailing_backslashes % 2:
            buffer = buffer[:-1]
            continue
        yield start_line, buffer
        buffer = ""
    if buffer:
        yield start_line, buffer


def _split_property(line: str) -> tuple[str, str] | None:
    stripped = line.lstrip(" \t\f")
    if not stripped or stripped[0] in "#!":
        return None

    escaped = False
    separator_index = len(stripped)
    for index, character in enumerate(stripped):
        if escaped:
            escaped = False
            continue
        if character == "\\":
            escaped = True
            continue
        if character in "=: \t\f":
            separator_index = index
            break

    value_index = separator_index
    while value_index < len(stripped) and stripped[value_index] in " \t\f":
        value_index += 1
    if value_index < len(stripped) and stripped[value_index] in "=:":
        value_index += 1
    while value_index < len(stripped) and stripped[value_index] in " \t\f":
        value_index += 1
    return stripped[:separator_index], stripped[value_index:]


def _unescape(value: str, line_number: int) -> str:
    result: list[str] = []
    index = 0
    escapes = {"t": "\t", "n": "\n", "r": "\r", "f": "\f"}
    while index < len(value):
        character = value[index]
        if character != "\\":
            result.append(character)
            index += 1
            continue
        index += 1
        if index >= len(value):
            result.append("\\")
            break
        escaped = value[index]
        if escaped == "u":
            codepoint = value[index + 1:index + 5]
            if len(codepoint) != 4 or not all(character in "0123456789abcdefABCDEF" for character in codepoint):
                raise PropertiesFormatError(f"Escape Unicode non valido alla riga {line_number}.")
            result.append(chr(int(codepoint, 16)))
            index += 5
            continue
        result.append(escapes.get(escaped, escaped))
        index += 1
    return "".join(result)


def _escape(value: str, *, key: bool) -> str:
    escaped: list[str] = []
    replacements = {"\\": "\\\\", "\t": "\\t", "\n": "\\n", "\r": "\\r", "\f": "\\f"}
    for index, character in enumerate(value):
        if character in replacements:
            escaped.append(replacements[character])
        elif key and character in " =:#!":
            escaped.append(f"\\{character}")
        elif not key and index == 0 and character in " #!":
            escaped.append(f"\\{character}")
        else:
            escaped.append(character)
    return "".join(escaped)


def read_properties_file(filepath: str | os.PathLike[str]) -> dict[str, str]:
    """Legge un file UTF-8 ``.properties`` senza perdere escape o continuazioni."""
    data: dict[str, str] = {}
    with open(filepath, encoding="utf-8-sig") as stream:
        for line_number, logical_line in _logical_lines(stream):
            pair = _split_property(logical_line)
            if pair is None:
                continue
            key = _unescape(pair[0], line_number)
            value = _unescape(pair[1], line_number)
            if key in data:
                raise PropertiesFormatError(f"Chiave duplicata '{key}' alla riga {line_number} di {filepath}.")
            data[key] = value
    return data


def _language_code(filepath: str | os.PathLike[str]) -> str:
    stem = Path(filepath).stem
    code = stem.rsplit("_", 1)[-1]
    if not _LANGUAGE_PATTERN.fullmatch(code):
        raise PropertiesFormatError(f"Codice lingua non valido nel file '{Path(filepath).name}'.")
    return code


def export_to_excel(properties_files: Iterable[str | os.PathLike[str]], output_path: str | os.PathLike[str]) -> None:
    """Converte più file ``.properties`` in un workbook con una colonna per lingua."""
    files = list(properties_files)
    if not files:
        raise PropertiesFormatError("Selezionare almeno un file .properties.")

    all_data: dict[str, dict[str, str]] = {}
    languages: set[str] = set()
    for filepath in files:
        language = _language_code(filepath)
        normalized_language = language.casefold()
        if normalized_language in languages:
            raise PropertiesFormatError(f"Codice lingua duplicato: '{language}'.")
        languages.add(normalized_language)
        for key, value in read_properties_file(filepath).items():
            all_data.setdefault(key, {})[language] = value

    dataframe = pd.DataFrame.from_dict(all_data, orient="index").rename_axis("key").reset_index()
    dataframe = dataframe.sort_values(by="key", key=lambda values: values.str.casefold())
    dataframe.to_excel(output_path, index=False, engine="openpyxl")

    # openpyxl interpreta automaticamente le stringhe che iniziano con "=" come formule.
    # Le traduzioni sono dati, non codice Excel: forza quindi ogni cella testuale al tipo stringa.
    workbook = load_workbook(output_path, data_only=False)
    worksheet = workbook.active
    for row in worksheet.iter_rows():
        for cell in row:
            if isinstance(cell.value, str):
                cell.data_type = "s"
    workbook.save(output_path)


def import_from_excel(excel_file: str | os.PathLike[str], output_dir: str | os.PathLike[str]) -> None:
    """Rigenera atomicamente un file UTF-8 ``.properties`` per ogni colonna lingua."""
    workbook = load_workbook(excel_file, data_only=False, read_only=True)
    try:
        formulas = [cell.coordinate for row in workbook.active.iter_rows() for cell in row if cell.data_type == "f"]
    finally:
        workbook.close()
    if formulas:
        preview = ", ".join(formulas[:5])
        suffix = "…" if len(formulas) > 5 else ""
        raise PropertiesFormatError(
            f"Il workbook contiene formule ({preview}{suffix}). Sostituirle con testo prima della conversione."
        )

    dataframe = pd.read_excel(excel_file, dtype=object, engine="openpyxl")
    key_columns = [column for column in dataframe.columns if str(column).casefold() == "key"]
    if len(key_columns) != 1:
        raise PropertiesFormatError("Il workbook deve contenere una sola colonna 'key'.")
    key_column = key_columns[0]
    languages = [str(column) for column in dataframe.columns if column != key_column]
    if not languages:
        raise PropertiesFormatError("Il workbook non contiene colonne lingua.")
    invalid_languages = [language for language in languages if not _LANGUAGE_PATTERN.fullmatch(language)]
    if invalid_languages:
        raise PropertiesFormatError(f"Codici lingua non validi: {', '.join(invalid_languages)}.")

    keys = [str(value) for value in dataframe[key_column] if pd.notna(value)]
    duplicates = sorted({key for key in keys if keys.count(key) > 1})
    if duplicates:
        raise PropertiesFormatError(f"Chiavi duplicate nel workbook: {', '.join(duplicates)}.")
    if len(keys) != len(dataframe.index):
        raise PropertiesFormatError("La colonna 'key' contiene celle vuote.")

    destination = Path(output_dir)
    destination.mkdir(parents=True, exist_ok=True)
    for language in languages:
        target = destination / f"Language_{language}.properties"
        temporary_path: str | None = None
        try:
            with tempfile.NamedTemporaryFile("w", encoding="utf-8", newline="\n", dir=destination, delete=False) as temporary:
                temporary_path = temporary.name
                for _, row in dataframe.iterrows():
                    value = row[language]
                    if pd.notna(value):
                        temporary.write(f"{_escape(str(row[key_column]), key=True)}={_escape(str(value), key=False)}\n")
            os.replace(temporary_path, target)
        finally:
            if temporary_path and os.path.exists(temporary_path):
                os.unlink(temporary_path)
